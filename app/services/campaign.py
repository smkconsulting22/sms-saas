import csv
import io
import asyncio
import logging
from typing import Optional
import phonenumbers
import openpyxl
from sqlalchemy.orm import Session
from app.models.contact import Contact
from app.models.campaign import Campaign, CampaignLog, CampaignStatus
from app.models.credit import CreditBalance
from app.models.user import User
from app.models.tenant import Tenant
from app.services.orange_sms import send_sms

logger = logging.getLogger(__name__)
LOW_BALANCE_THRESHOLD = 10


def validate_phone(phone: str) -> Optional[str]:
    """Retourne le numéro formaté E.164 ou None si invalide"""
    try:
        # Nettoyer le numéro
        cleaned = phone.strip().replace(" ", "").replace("-", "").replace(".", "")
        print(f"N°1: {cleaned}")
        # Si le numéro commence par 0 (format local ivoirien ex: 0502787818)
        if cleaned.startswith("0") and len(cleaned) == 10:
            cleaned = "225" + cleaned  # 0502787818 → 2250502787818

        # Ajouter le + si absent
        if not cleaned.startswith("+"):
            cleaned = "+" + cleaned
        print(f"N°2: {cleaned}")
        parsed = phonenumbers.parse(cleaned, "CI")
        if phonenumbers.is_valid_number(parsed):
            return phonenumbers.format_number(
                parsed, phonenumbers.PhoneNumberFormat.E164
            )
    except Exception:
        pass
    return None

def import_contacts_from_csv(
    db: Session,
    tenant_id: str,
    content: bytes
) -> dict:
    """Importe des contacts depuis un fichier CSV (UTF-8 ou Latin-1)"""

    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = content.decode("latin-1")
        except UnicodeDecodeError:
            decoded = content.decode("cp1252")

    sample = decoded[:1024]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)

    imported = 0
    skipped = 0
    duplicates = 0
    errors = []

    for i, row in enumerate(reader, start=2):
        normalized = {k.strip().lower(): v.strip() for k, v in row.items() if k}

        phone_raw = (
            normalized.get("phone") or
            normalized.get("telephone") or
            normalized.get("téléphone") or
            normalized.get("numero") or
            normalized.get("numéro") or
            normalized.get("tel") or
            normalized.get("mobile") or
            ""
        )

        if not phone_raw:
            errors.append(f"Ligne {i} : colonne téléphone introuvable")
            skipped += 1
            continue

        phone = validate_phone(phone_raw)
        if not phone:
            errors.append(f"Ligne {i} : numéro invalide '{phone_raw}'")
            skipped += 1
            continue

        full_name = (
            normalized.get("full_name") or
            normalized.get("nom") or
            normalized.get("name") or
            normalized.get("prenom") or
            normalized.get("prénom") or
            ""
        )

        existing = db.query(Contact).filter(
            Contact.tenant_id == tenant_id,
            Contact.phone == phone
        ).first()

        if existing:
            duplicates += 1
            continue

        contact = Contact(
            tenant_id=tenant_id,
            phone=phone,
            full_name=full_name.strip()
        )
        db.add(contact)
        imported += 1

    db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "duplicates": duplicates,
        "errors": errors[:10]
    }


def import_contacts_from_excel(
    db: Session,
    tenant_id: str,
    content: bytes
) -> dict:
    """Importe des contacts depuis un fichier Excel .xlsx"""

    wb = openpyxl.load_workbook(filename=io.BytesIO(content))
    ws = wb.active

    imported = 0
    skipped = 0
    duplicates = 0
    errors = []

    # Lire la première ligne comme en-têtes
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower()
        headers.append(val)

    # Trouver les index des colonnes
    phone_col = next((i for i, h in enumerate(headers) if h in [
        "phone", "telephone", "téléphone", "numero", "numéro", "tel", "mobile"
    ]), None)

    name_col = next((i for i, h in enumerate(headers) if h in [
        "full_name", "nom", "name", "prenom", "prénom"
    ]), None)

    if phone_col is None:
        return {
            "imported": 0,
            "skipped": 0,
            "duplicates": 0,
            "errors": ["Colonne téléphone introuvable. Colonnes attendues : phone, telephone, numero, tel, mobile"]
        }

    # Parcourir les lignes à partir de la ligne 2
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        phone_raw = str(row[phone_col] or "").strip()
        if not phone_raw or phone_raw == "None":
            errors.append(f"Ligne {i} : numéro vide")
            skipped += 1
            continue

        phone = validate_phone(phone_raw)
        if not phone:
            errors.append(f"Ligne {i} : numéro invalide '{phone_raw}'")
            skipped += 1
            continue
        
        full_name = ""
        if name_col is not None and row[name_col]:
            full_name = str(row[name_col]).strip()

        # Vérifier les doublons
        existing = db.query(Contact).filter(
            Contact.tenant_id == tenant_id,
            Contact.phone == phone
        ).first()

        if existing:
            duplicates += 1
            continue

        contact = Contact(
            tenant_id=tenant_id,
            phone=phone,
            full_name=full_name
        )
        db.add(contact)
        imported += 1

    db.commit()

    # Log temporaire
    print(f"HEADERS: {headers}")
    print(f"PHONE COL INDEX: {phone_col}")
    print(f"NAME COL INDEX: {name_col}")

    # Afficher les 3 premières lignes
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        print(f"LIGNE {i+2}: {row}")

    return {
        "imported": imported,
        "skipped": skipped,
        "duplicates": duplicates,
        "errors": errors[:10]
    }

async def run_campaign(db: Session, campaign_id: str, tenant_id: str):
    """Envoie les SMS à tous les contacts actifs du tenant"""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        return

    contacts = db.query(Contact).filter(
        Contact.tenant_id == tenant_id,
        Contact.is_optout == False
    ).all()

    balance = db.query(CreditBalance).filter(
        CreditBalance.tenant_id == tenant_id
    ).first()

    if not balance or balance.balance < len(contacts):
        campaign.status = CampaignStatus.FAILED
        db.commit()
        return

    # Récupérer le sender_name personnalisé du tenant (None → fallback "RESAFIG")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    sender_name = tenant.sender_name if tenant else None

    campaign.status = CampaignStatus.RUNNING
    campaign.total = len(contacts)
    db.commit()

    for contact in contacts:
        message = campaign.message.replace(
            "{{nom}}", contact.full_name or "Client"
        )

        try:
            result = await send_sms(contact.phone, message, tenant_id=tenant_id, sender_name=sender_name)
            resource_url = result.get(
                "outboundSMSMessageRequest", {}
            ).get("resourceURL", "")
            message_id = resource_url.split("/")[-1] if resource_url else None

            log = CampaignLog(
                campaign_id=campaign.id,
                phone=contact.phone,
                status="sent",
                message_id=message_id
            )
            campaign.sent += 1
            balance.balance -= 1

        except Exception as e:
            log = CampaignLog(
                campaign_id=campaign.id,
                phone=contact.phone,
                status="failed",
                error=str(e)
            )
            campaign.failed += 1

        db.add(log)
        db.commit()
        await asyncio.sleep(0.1)  # 100ms entre chaque SMS

    campaign.status = CampaignStatus.COMPLETED
    db.commit()

    # Alerte solde bas si balance < seuil — non bloquant
    if balance and balance.balance < LOW_BALANCE_THRESHOLD:
        try:
            from app.services.email import send_low_balance_alert
            tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
            admin = db.query(User).filter(
                User.tenant_id == tenant_id,
                User.role == "admin",
                User.is_active == True,
            ).first()
            if tenant and admin:
                await asyncio.to_thread(
                    send_low_balance_alert, admin.email, tenant.name, balance.balance
                )
        except Exception as exc:
            logger.error("Erreur alerte solde bas : %s", exc)